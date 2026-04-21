"""
ExcelExporter Service

Generates a multi-tab Excel workbook from normalized PostgreSQL database.
Each sheet acts as a relational table, with Candidate_ID and Candidate_Name
as the first two columns for cross-referencing across sheets.

Product: TALASH v3 Smart HR Recruitment System
Author: Data Engineering Team
"""

import os
from typing import Dict, List
from datetime import datetime, date

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models.models import (
    Candidate,
    EducationRecord,
    WorkExperience,
    JournalPublication,
    ConferencePublication,
    SupervisionRecord,
    Book,
    Patent,
    Skill,
    ExtractionRun,
    PublicationAuthor,
    EducationGap,
    EmploymentGap,
    InstitutionRanking,
    TopicCluster,
    CollaborationEdge,
    CandidateAssessment,
    MissingInformationRequest,
)


class ExcelExporter:
    """
    Generates a production-ready multi-tab Excel workbook from the CV extraction database.
    
    Each sheet represents a relational table with:
    - Candidate_ID and Candidate_Name as the first two columns (for cross-referencing)
    - Formatted headers (bold, light gray background)
    - Frozen top row
    - Auto-adjusted column widths (max 50 chars)
    - Auto-filtering enabled
    """

    def __init__(self, db: Session):
        """
        Initialize the ExcelExporter.
        
        Args:
            db: SQLAlchemy database session
        """
        self.db = db
        self.header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        self.header_font = Font(bold=True)

    def generate_master_report(self, output_path: str = "backend/data/exports/Master_CV_Report.xlsx") -> str:
        """
        Generate the master multi-tab Excel report.
        
        Args:
            output_path: Path to save the Excel file (default: backend/data/exports/Master_CV_Report.xlsx)
        
        Returns:
            str: Path to the generated file
        """
        # Ensure output directory exists
        output_dir = os.path.dirname(output_path)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)

        # Dictionary to hold all sheets
        sheets_dict = {
            "Candidates": self._get_candidates_df(),
            "Education": self._get_education_df(),
            "Experience": self._get_experience_df(),
            "Journals": self._get_journals_df(),
            "Conferences": self._get_conferences_df(),
            "Supervision": self._get_supervision_df(),
            "Books": self._get_books_df(),
            "Patents": self._get_patents_df(),
            "Skills": self._get_skills_df(),
            "ExtractionRuns": self._get_extraction_runs_df(),
            "PublicationAuthors": self._get_publication_authors_df(),
            "EducationGaps": self._get_education_gaps_df(),
            "EmploymentGaps": self._get_employment_gaps_df(),
            "InstitutionRankings": self._get_institution_rankings_df(),
            "TopicClusters": self._get_topic_clusters_df(),
            "CollaborationEdges": self._get_collaboration_edges_df(),
            "Assessments": self._get_assessments_df(),
            "MissingInfoRequests": self._get_missing_info_requests_df(),
        }

        # Write to Excel with openpyxl engine for formatting
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            for sheet_name, df in sheets_dict.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Apply formatting to all sheets
        self._apply_formatting(output_path)

        return output_path

    # ============================================================================
    # DATA FETCHING METHODS (Modular functions for each sheet)
    # ============================================================================

    def _get_candidates_df(self) -> pd.DataFrame:
        """
        Fetch candidates data.
        
        Returns:
            DataFrame with columns: [Candidate_ID, Candidate_Name, Email, Phone, LinkedIn, Profile_Summary, Total_Score]
        """
        candidates = self.db.query(Candidate).all()
        
        data = []
        for candidate in candidates:
            data.append({
                "Candidate_ID": candidate.id,
                "Candidate_Name": candidate.name,
                "Email": candidate.email or "",
                "Phone": candidate.phone or "",
                "LinkedIn": candidate.linkedin_url or "",
                "Personal_Website": candidate.personal_website or "",
                "Other_URLs": candidate.other_urls or "",
                "Profile_Summary": (candidate.summary[:100] + "...") if candidate.summary and len(candidate.summary) > 100 else (candidate.summary or ""),
                "Status": candidate.status or "",
                "Total_Score": "",
            })

        df = pd.DataFrame(data)
        
        # Return empty DataFrame with correct columns if no data
        if df.empty:
            return pd.DataFrame(columns=[
                "Candidate_ID", "Candidate_Name", "Email", "Phone", "LinkedIn", "Personal_Website", "Other_URLs", "Profile_Summary", "Status", "Total_Score"
            ])
        
        return df

    def _get_education_df(self) -> pd.DataFrame:
        """
        Fetch education records.
        
        Returns:
            DataFrame with education records for all candidates.
        """
        education_records = self.db.query(EducationRecord).all()
        
        data = []
        for edu in education_records:
            candidate = self.db.query(Candidate).filter(Candidate.id == edu.candidate_id).first()
            data.append({
                "Candidate_ID": edu.candidate_id,
                "Candidate_Name": candidate.name if candidate else "",
                "Institution_Type": edu.institution_type or "",
                "Degree_Level": edu.stage or "",
                "Degree_Title": edu.degree_title or "",
                "Institution": edu.institution or "",
                "Board_or_University": edu.board_or_university or "",
                "Start_Year": edu.start_year or "",
                "Start_Month": edu.start_month or "",
                "End_Year": edu.end_year or "",
                "End_Month": edu.end_month or "",
                "Marks_Percentage": edu.marks_percentage or "",
                "CGPA": edu.cgpa or "",
                "CGPA_Scale": edu.cgpa_scale or "",
                "Normalized_CGPA": edu.normalized_cgpa or "",
                "THE_Ranking": edu.institution_the_ranking or "",
                "QS_Ranking": edu.institution_qs_ranking or "",
                "Ranking_Source": edu.institution_ranking_source or "",
                "Ranking_Year": edu.institution_ranking_year or "",
                "Ranking_Value": edu.institution_ranking_value or "",
                "Gap_Before_Start_Months": edu.gap_before_start_months or "",
                "Gap_Justified_By_Experience": "Yes" if edu.gap_justified_by_experience else "No",
                "Evidence_Text": edu.evidence_text or "",
                "Confidence_Score": edu.confidence_score or "",
            })

        df = pd.DataFrame(data)
        
        if df.empty:
            return pd.DataFrame(columns=[
                "Candidate_ID", "Candidate_Name", "Institution_Type", "Degree_Level", "Degree_Title",
                "Institution", "Board_or_University", "Start_Year", "Start_Month", "End_Year", "End_Month",
                "Marks_Percentage", "CGPA", "CGPA_Scale", "Normalized_CGPA", "THE_Ranking", "QS_Ranking",
                "Ranking_Source", "Ranking_Year", "Ranking_Value", "Gap_Before_Start_Months",
                "Gap_Justified_By_Experience", "Evidence_Text", "Confidence_Score"
            ])
        
        return df

    def _get_experience_df(self) -> pd.DataFrame:
        """
        Fetch work experience records.
        
        Returns:
            DataFrame with work experience for all candidates.
        """
        work_experiences = self.db.query(WorkExperience).all()
        
        data = []
        for exp in work_experiences:
            candidate = self.db.query(Candidate).filter(Candidate.id == exp.candidate_id).first()
            data.append({
                "Candidate_ID": exp.candidate_id,
                "Candidate_Name": candidate.name if candidate else "",
                "Job_Title": exp.job_title or "",
                "Organization": exp.organization or "",
                "Location": exp.location or "",
                "Employment_Type": exp.employment_type or "",
                "Start_Date": exp.start_date.strftime("%Y-%m-%d") if exp.start_date else "",
                "End_Date": exp.end_date.strftime("%Y-%m-%d") if exp.end_date else "",
                "Start_Month": exp.start_month or "",
                "Start_Year": exp.start_year or "",
                "End_Month": exp.end_month or "",
                "End_Year": exp.end_year or "",
                "Is_Current": "Yes" if exp.is_current else "No",
                "Academic_Role": "Yes" if exp.is_academic_role else "No",
                "Overlaps_With_Education": "Yes" if exp.overlaps_with_education else "No",
                "Job_Responsibilities": exp.job_responsibilities or "",
                "Evidence_Text": exp.evidence_text or "",
                "Confidence_Score": exp.confidence_score or "",
            })

        df = pd.DataFrame(data)
        
        if df.empty:
            return pd.DataFrame(columns=[
                "Candidate_ID", "Candidate_Name", "Job_Title", "Organization",
                "Location", "Employment_Type", "Start_Date", "End_Date", "Start_Month", "Start_Year",
                "End_Month", "End_Year", "Is_Current", "Academic_Role", "Overlaps_With_Education",
                "Job_Responsibilities", "Evidence_Text", "Confidence_Score"
            ])
        
        return df

    def _get_journals_df(self) -> pd.DataFrame:
        """
        Fetch journal publications (type='journal').
        
        Returns:
            DataFrame with journal publications.
        """
        publications = self.db.query(JournalPublication).all()
        
        data = []
        for pub in publications:
            candidate = self.db.query(Candidate).filter(Candidate.id == pub.candidate_id).first()
            data.append({
                "Candidate_ID": pub.candidate_id,
                "Candidate_Name": candidate.name if candidate else "",
                "Title": pub.title or "",
                "Journal_Name": pub.journal_name or "",
                "ISSN": pub.issn or "",
                "DOI": pub.doi or "",
                "Year": pub.year or "",
                "Volume": pub.volume or "",
                "Issue": pub.issue or "",
                "Pages": pub.pages or "",
                "Authorship_Role": pub.authorship_role or "Co-Author",
                "WoS_Indexed": "Yes" if pub.wos_indexed else "No",
                "Scopus_Indexed": "Yes" if pub.scopus_indexed else "No",
                "Quartile": pub.quartile or "Unknown",
                "Impact_Factor": pub.impact_factor or "",
                "Topic_Category": pub.topic_category or "",
                "Is_With_Student": "Yes" if pub.is_with_student else "No",
                "Abstract_or_Summary": pub.abstract_or_summary or "",
                "Keywords_JSON": pub.keywords_json or "",
                "Author_Affiliations_JSON": pub.author_affiliations_json or "",
                "Source_Verification_URL": pub.source_verification_url or "",
                "Confidence_Score": pub.confidence_score or "",
            })

        df = pd.DataFrame(data)
        
        if df.empty:
            return pd.DataFrame(columns=[
                "Candidate_ID", "Candidate_Name", "Title", "Journal_Name", "ISSN", "DOI",
                "Year", "Volume", "Issue", "Pages", "Authorship_Role", "WoS_Indexed", "Scopus_Indexed",
                "Quartile", "Impact_Factor", "Topic_Category", "Is_With_Student", "Abstract_or_Summary",
                "Keywords_JSON", "Author_Affiliations_JSON", "Source_Verification_URL", "Confidence_Score"
            ])
        
        return df

    def _get_conferences_df(self) -> pd.DataFrame:
        """
        Fetch conference publications (type='conference').
        
        Returns:
            DataFrame with conference publications.
        """
        publications = self.db.query(ConferencePublication).all()
        
        data = []
        for pub in publications:
            candidate = self.db.query(Candidate).filter(Candidate.id == pub.candidate_id).first()
            data.append({
                "Candidate_ID": pub.candidate_id,
                "Candidate_Name": candidate.name if candidate else "",
                "Title": pub.title or "",
                "Conference_Name": pub.conference_name or "",
                "Conference_Series": pub.conference_series or "",
                "Conference_Location": pub.conference_location or "",
                "Year": pub.year or "",
                "CORE_Ranking": pub.core_ranking or "Unranked",
                "A_Star_Ranked": "Yes" if pub.is_a_star else "No",
                "Authorship_Role": pub.authorship_role or "Co-Author",
                "Indexed_In": pub.indexed_in or "",
                "Publisher": pub.publisher or "",
                "DOI": pub.doi or "",
                "Topic_Category": pub.topic_category or "",
                "Is_With_Student": "Yes" if pub.is_with_student else "No",
                "Abstract_or_Summary": pub.abstract_or_summary or "",
                "Keywords_JSON": pub.keywords_json or "",
                "Author_Affiliations_JSON": pub.author_affiliations_json or "",
                "Source_Verification_URL": pub.source_verification_url or "",
                "Confidence_Score": pub.confidence_score or "",
            })

        df = pd.DataFrame(data)
        
        if df.empty:
            return pd.DataFrame(columns=[
                "Candidate_ID", "Candidate_Name", "Title", "Conference_Name",
                "Conference_Series", "Conference_Location", "Year", "CORE_Ranking", "A_Star_Ranked",
                "Authorship_Role", "Indexed_In", "Publisher", "DOI", "Topic_Category",
                "Is_With_Student", "Abstract_or_Summary", "Keywords_JSON", "Author_Affiliations_JSON",
                "Source_Verification_URL", "Confidence_Score"
            ])
        
        return df

    def _get_supervision_df(self) -> pd.DataFrame:
        """
        Fetch supervision records.
        
        Returns:
            DataFrame with supervision records.
        """
        supervision_records = self.db.query(SupervisionRecord).all()
        
        data = []
        for sup in supervision_records:
            candidate = self.db.query(Candidate).filter(Candidate.id == sup.candidate_id).first()
            data.append({
                "Candidate_ID": sup.candidate_id,
                "Candidate_Name": candidate.name if candidate else "",
                "Student_Name": sup.student_name or "",
                "Student_Level": sup.student_level or "",
                "Year": sup.completion_year or "",
                "Supervision_Role": sup.supervision_role or "",
                "Publications_With_Student": sup.publications_with_student or 0,
                "Thesis_Title": sup.thesis_title or "",
                "Evidence_Text": sup.evidence_text or "",
                "Confidence_Score": sup.confidence_score or "",
            })

        df = pd.DataFrame(data)
        
        if df.empty:
            return pd.DataFrame(columns=[
                "Candidate_ID", "Candidate_Name", "Student_Name", "Student_Level", "Year",
                "Supervision_Role", "Publications_With_Student", "Thesis_Title", "Evidence_Text", "Confidence_Score"
            ])
        
        return df

    def _get_books_df(self) -> pd.DataFrame:
        """
        Fetch book records.
        
        Returns:
            DataFrame with books.
        """
        books = self.db.query(Book).all()
        
        data = []
        for book in books:
            candidate = self.db.query(Candidate).filter(Candidate.id == book.candidate_id).first()
            data.append({
                "Candidate_ID": book.candidate_id,
                "Candidate_Name": candidate.name if candidate else "",
                "Title": book.title or "",
                "Authors": book.authors or "",
                "Publisher": book.publisher or "",
                "Year": book.year or "",
                "ISBN": book.isbn or "",
                "Authorship_Role": book.authorship_role or "",
                "Online_Link": book.online_link or "",
                "Evidence_Text": book.evidence_text or "",
                "Confidence_Score": book.confidence_score or "",
            })

        df = pd.DataFrame(data)
        
        if df.empty:
            return pd.DataFrame(columns=[
                "Candidate_ID", "Candidate_Name", "Title", "Authors", "Publisher", "Year", "ISBN",
                "Authorship_Role", "Online_Link", "Evidence_Text", "Confidence_Score"
            ])
        
        return df

    def _get_patents_df(self) -> pd.DataFrame:
        """
        Fetch patent records.
        
        Returns:
            DataFrame with patents.
        """
        patents = self.db.query(Patent).all()
        
        data = []
        for patent in patents:
            candidate = self.db.query(Candidate).filter(Candidate.id == patent.candidate_id).first()
            data.append({
                "Candidate_ID": patent.candidate_id,
                "Candidate_Name": candidate.name if candidate else "",
                "Title": patent.title or "",
                "Patent_No": patent.patent_no or "",
                "Inventors": patent.inventors or "",
                "Year": patent.date_filed.year if patent.date_filed else "",
                "Date_Granted": patent.date_granted.strftime("%Y-%m-%d") if patent.date_granted else "",
                "Country_Of_Filing": patent.country_of_filing or "",
                "Online_Link": patent.online_link or "",
                "Inventor_Role": patent.inventor_role or "",
                "Status": patent.status or "",
                "Evidence_Text": patent.evidence_text or "",
                "Confidence_Score": patent.confidence_score or "",
            })

        df = pd.DataFrame(data)
        
        if df.empty:
            return pd.DataFrame(columns=[
                "Candidate_ID", "Candidate_Name", "Title", "Patent_No", "Inventors", "Year", "Date_Granted",
                "Country_Of_Filing", "Online_Link", "Inventor_Role", "Status", "Evidence_Text", "Confidence_Score"
            ])
        
        return df

    def _get_skills_df(self) -> pd.DataFrame:
        """
        Fetch skill records.
        
        Returns:
            DataFrame with skills.
        """
        skills = self.db.query(Skill).all()
        
        data = []
        for skill in skills:
            candidate = self.db.query(Candidate).filter(Candidate.id == skill.candidate_id).first()
            data.append({
                "Candidate_ID": skill.candidate_id,
                "Candidate_Name": candidate.name if candidate else "",
                "Skill_Name": skill.name or "",
                "Category": skill.category or "",
                "Proficiency_Level": skill.proficiency_level or "",
                "Years_of_Experience": skill.years_of_experience or "",
                "Evidenced_In_Work": "Yes" if skill.evidenced_in_work else "No",
                "Evidenced_In_Research": "Yes" if skill.evidenced_in_research else "No",
                "Work_Evidence": skill.work_evidence or "",
                "Research_Evidence": skill.research_evidence or "",
                "Strength_Of_Evidence": skill.strength_of_evidence or "",
                "Confidence_Score": skill.confidence_score or "",
            })

        df = pd.DataFrame(data)
        
        if df.empty:
            return pd.DataFrame(columns=[
                "Candidate_ID", "Candidate_Name", "Skill_Name", "Category", "Proficiency_Level",
                "Years_of_Experience", "Evidenced_In_Work", "Evidenced_In_Research", "Work_Evidence",
                "Research_Evidence", "Strength_Of_Evidence", "Confidence_Score"
            ])

        return df

    def _get_extraction_runs_df(self) -> pd.DataFrame:
        runs = self.db.query(ExtractionRun).all()
        data = []
        for run in runs:
            candidate = self.db.query(Candidate).filter(Candidate.id == run.candidate_id).first()
            data.append({
                "Candidate_ID": run.candidate_id,
                "Candidate_Name": candidate.name if candidate else "",
                "Provider": run.provider,
                "Model_Name": run.model_name,
                "Prompt_Version": run.prompt_version or "",
                "Run_Type": run.run_type,
                "Status": run.status,
                "Parsed_OK": "Yes" if run.parsed_ok else "No",
                "Error_Message": run.error_message or "",
                "Started_At": run.started_at.isoformat() if run.started_at else "",
                "Finished_At": run.finished_at.isoformat() if run.finished_at else "",
                "Raw_Response_JSON": run.raw_response_json or "",
            })
        if not data:
            return pd.DataFrame(columns=[
                "Candidate_ID", "Candidate_Name", "Provider", "Model_Name", "Prompt_Version", "Run_Type",
                "Status", "Parsed_OK", "Error_Message", "Started_At", "Finished_At", "Raw_Response_JSON"
            ])
        return pd.DataFrame(data)

    def _get_publication_authors_df(self) -> pd.DataFrame:
        authors = self.db.query(PublicationAuthor).all()
        data = []
        for author in authors:
            candidate = self.db.query(Candidate).filter(Candidate.id == author.candidate_id).first()
            data.append({
                "Candidate_ID": author.candidate_id,
                "Candidate_Name": candidate.name if candidate else "",
                "Publication_Type": author.publication_type,
                "Publication_ID": author.publication_id,
                "Author_Order": author.author_order or "",
                "Author_Name": author.author_name,
                "Is_Candidate": "Yes" if author.is_candidate else "No",
                "Is_Corresponding": "Yes" if author.is_corresponding else "No",
                "Affiliation": author.affiliation or "",
                "Normalized_Author_Key": author.normalized_author_key or "",
            })
        if not data:
            return pd.DataFrame(columns=[
                "Candidate_ID", "Candidate_Name", "Publication_Type", "Publication_ID", "Author_Order",
                "Author_Name", "Is_Candidate", "Is_Corresponding", "Affiliation", "Normalized_Author_Key"
            ])
        return pd.DataFrame(data)

    def _get_education_gaps_df(self) -> pd.DataFrame:
        gaps = self.db.query(EducationGap).all()
        data = []
        for gap in gaps:
            candidate = self.db.query(Candidate).filter(Candidate.id == gap.candidate_id).first()
            data.append({
                "Candidate_ID": gap.candidate_id,
                "Candidate_Name": candidate.name if candidate else "",
                "From_Stage": gap.from_stage or "",
                "To_Stage": gap.to_stage or "",
                "Gap_Months": gap.gap_months or "",
                "Gap_Start": gap.gap_start.isoformat() if gap.gap_start else "",
                "Gap_End": gap.gap_end.isoformat() if gap.gap_end else "",
                "Justified_By_Work": "Yes" if gap.justified_by_work else "No",
                "Justification_Text": gap.justification_text or "",
                "Evidence_Work_Experience_ID": gap.evidence_work_experience_id or "",
            })
        if not data:
            return pd.DataFrame(columns=[
                "Candidate_ID", "Candidate_Name", "From_Stage", "To_Stage", "Gap_Months", "Gap_Start",
                "Gap_End", "Justified_By_Work", "Justification_Text", "Evidence_Work_Experience_ID"
            ])
        return pd.DataFrame(data)

    def _get_employment_gaps_df(self) -> pd.DataFrame:
        gaps = self.db.query(EmploymentGap).all()
        data = []
        for gap in gaps:
            candidate = self.db.query(Candidate).filter(Candidate.id == gap.candidate_id).first()
            data.append({
                "Candidate_ID": gap.candidate_id,
                "Candidate_Name": candidate.name if candidate else "",
                "Gap_Type": gap.gap_type or "",
                "Gap_Start": gap.gap_start.isoformat() if gap.gap_start else "",
                "Gap_End": gap.gap_end.isoformat() if gap.gap_end else "",
                "Gap_Months": gap.gap_months or "",
                "Justified": "Yes" if gap.justified else "No",
                "Justification_Text": gap.justification_text or "",
                "Related_Education_ID": gap.related_education_id or "",
                "Related_Career_Break_ID": gap.related_career_break_id or "",
            })
        if not data:
            return pd.DataFrame(columns=[
                "Candidate_ID", "Candidate_Name", "Gap_Type", "Gap_Start", "Gap_End", "Gap_Months",
                "Justified", "Justification_Text", "Related_Education_ID", "Related_Career_Break_ID"
            ])
        return pd.DataFrame(data)

    def _get_institution_rankings_df(self) -> pd.DataFrame:
        rankings = self.db.query(InstitutionRanking).all()
        data = []
        for ranking in rankings:
            candidate = self.db.query(Candidate).filter(Candidate.id == ranking.candidate_id).first()
            data.append({
                "Candidate_ID": ranking.candidate_id,
                "Candidate_Name": candidate.name if candidate else "",
                "Institution_Name": ranking.institution_name,
                "Institution_Type": ranking.institution_type or "",
                "Source": ranking.source,
                "Source_Year": ranking.source_year or "",
                "Rank_Value": ranking.rank_value or "",
                "Rank_Band": ranking.rank_band or "",
                "Country": ranking.country or "",
                "URL": ranking.url or "",
                "Verified_At": ranking.verified_at.isoformat() if ranking.verified_at else "",
            })
        if not data:
            return pd.DataFrame(columns=[
                "Candidate_ID", "Candidate_Name", "Institution_Name", "Institution_Type", "Source",
                "Source_Year", "Rank_Value", "Rank_Band", "Country", "URL", "Verified_At"
            ])
        return pd.DataFrame(data)

    def _get_topic_clusters_df(self) -> pd.DataFrame:
        clusters = self.db.query(TopicCluster).all()
        data = []
        for cluster in clusters:
            candidate = self.db.query(Candidate).filter(Candidate.id == cluster.candidate_id).first()
            data.append({
                "Candidate_ID": cluster.candidate_id,
                "Candidate_Name": candidate.name if candidate else "",
                "Publication_Type": cluster.publication_type,
                "Publication_ID": cluster.publication_id,
                "Cluster_Name": cluster.cluster_name,
                "Cluster_Score": cluster.cluster_score or "",
                "Assigned_By": cluster.assigned_by or "",
                "Model_Version": cluster.model_version or "",
            })
        if not data:
            return pd.DataFrame(columns=[
                "Candidate_ID", "Candidate_Name", "Publication_Type", "Publication_ID", "Cluster_Name",
                "Cluster_Score", "Assigned_By", "Model_Version"
            ])
        return pd.DataFrame(data)

    def _get_collaboration_edges_df(self) -> pd.DataFrame:
        edges = self.db.query(CollaborationEdge).all()
        data = []
        for edge in edges:
            candidate = self.db.query(Candidate).filter(Candidate.id == edge.candidate_id).first()
            data.append({
                "Candidate_ID": edge.candidate_id,
                "Candidate_Name": candidate.name if candidate else "",
                "Coauthor_Name": edge.coauthor_name,
                "Coauthor_Affiliation": edge.coauthor_affiliation or "",
                "Publication_Type": edge.publication_type,
                "Publication_ID": edge.publication_id,
                "Edge_Weight": edge.edge_weight or "",
                "Is_Recurring": "Yes" if edge.is_recurring else "No",
            })
        if not data:
            return pd.DataFrame(columns=[
                "Candidate_ID", "Candidate_Name", "Coauthor_Name", "Coauthor_Affiliation", "Publication_Type",
                "Publication_ID", "Edge_Weight", "Is_Recurring"
            ])
        return pd.DataFrame(data)

    def _get_assessments_df(self) -> pd.DataFrame:
        assessments = self.db.query(CandidateAssessment).all()
        data = []
        for assessment in assessments:
            candidate = self.db.query(Candidate).filter(Candidate.id == assessment.candidate_id).first()
            data.append({
                "Candidate_ID": assessment.candidate_id,
                "Candidate_Name": candidate.name if candidate else "",
                "Assessment_Version": assessment.assessment_version or "",
                "Education_Strength_Score": assessment.education_strength_score or "",
                "Research_Strength_Score": assessment.research_strength_score or "",
                "Experience_Strength_Score": assessment.experience_strength_score or "",
                "Skill_Alignment_Score": assessment.skill_alignment_score or "",
                "Overall_Rank": assessment.overall_rank or "",
                "Overall_Summary": assessment.overall_summary or "",
                "Missing_Sections_JSON": assessment.missing_sections_json or "",
                "Generated_At": assessment.generated_at.isoformat() if assessment.generated_at else "",
            })
        if not data:
            return pd.DataFrame(columns=[
                "Candidate_ID", "Candidate_Name", "Assessment_Version", "Education_Strength_Score",
                "Research_Strength_Score", "Experience_Strength_Score", "Skill_Alignment_Score",
                "Overall_Rank", "Overall_Summary", "Missing_Sections_JSON", "Generated_At"
            ])
        return pd.DataFrame(data)

    def _get_missing_info_requests_df(self) -> pd.DataFrame:
        requests = self.db.query(MissingInformationRequest).all()
        data = []
        for request in requests:
            candidate = self.db.query(Candidate).filter(Candidate.id == request.candidate_id).first()
            data.append({
                "Candidate_ID": request.candidate_id,
                "Candidate_Name": candidate.name if candidate else "",
                "Module_Name": request.module_name,
                "Missing_Fields_JSON": request.missing_fields_json or "",
                "Draft_Email_Subject": request.draft_email_subject or "",
                "Draft_Email_Body": request.draft_email_body or "",
                "Status": request.status,
                "Generated_At": request.generated_at.isoformat() if request.generated_at else "",
                "Sent_At": request.sent_at.isoformat() if request.sent_at else "",
            })
        if not data:
            return pd.DataFrame(columns=[
                "Candidate_ID", "Candidate_Name", "Module_Name", "Missing_Fields_JSON", "Draft_Email_Subject",
                "Draft_Email_Body", "Status", "Generated_At", "Sent_At"
            ])
        return pd.DataFrame(data)

    # ============================================================================
    # FORMATTING METHODS
    # ============================================================================

    def _apply_formatting(self, output_path: str) -> None:
        """
        Apply professional formatting to all sheets:
        - Freeze top row
        - Bold header with light gray background
        - Auto-adjusted column widths (max 50)
        - Auto-filtering on headers
        
        Args:
            output_path: Path to the Excel file
        """
        workbook = load_workbook(output_path)
        
        for sheet in workbook.sheetnames:
            ws = workbook[sheet]
            
            # Freeze the top row
            ws.freeze_panes = "A2"
            
            # Format header row (row 1)
            for cell in ws[1]:
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Auto-adjust column widths
            for col_num, column_cells in enumerate(ws.columns, 1):
                max_length = 0
                column_letter = get_column_letter(col_num)
                
                for cell in column_cells:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except Exception:
                        pass
                
                # Set width with a maximum of 50 characters
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Enable auto-filtering on the header row
            if ws.dimensions:
                ws.auto_filter.ref = ws.dimensions
        
        workbook.save(output_path)
